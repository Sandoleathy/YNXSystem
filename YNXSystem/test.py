from openpyxl import load_workbook

wp = load_workbook("YNXSystem/algorithm/algorithm.xlsx")
sheet = wp.active

for row in sheet.iter_rows(values_only=True):
    print(row[0])