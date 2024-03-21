from YNXSystem import app
from flask import render_template, Flask, request, jsonify
from openpyxl import load_workbook #用于处理excel的库
from datetime import datetime

@app.route('/', methods=['POST','GET'])
@app.route('/index', methods=['POST','GET'])
def index():
    return render_template("index.html")
@app.route('/calculate', methods=['POST'])
def calculate():
    data = request.form["date"]
    timeSet = data.split("-")
    #切分时间字符串
    print(timeSet)
    year = int(timeSet[0])
    month = int(timeSet[1])
    day = int(timeSet[2])

    yearNo = year%6
    queryYear = 0
    if(yearNo == 1):
        queryYear = 2023
    elif(yearNo == 2):
        queryYear = 2024
    elif(yearNo == 3):
        queryYear = 2025
    elif(yearNo == 4):
        queryYear = 2026
    elif(yearNo == 5):
        queryYear = 2027
    elif(yearNo == 0):
        queryYear = 2028
    queryTime = datetime(queryYear,month,day)

    algorithm_worksheet = load_workbook('YNXSystem/algorithm/algorithm.xlsx')
    sheet = algorithm_worksheet.active
    counter = 0
    isGetResult = False
    for row in sheet.iter_rows(values_only=True):
        timeStr = row[0].split("-")
        currentTime = datetime(int(timeStr[0]),int(timeStr[1]),int(timeStr[2]))
        #print(currentTime)
        #print(queryTime)
        if(counter == 0 and queryTime < currentTime):
            queryTime = (2028,month,day)
        if(queryTime > currentTime):
            pass
        else:
            result = row[1]
            monthExp = row[2]
            yearExp = row[3]
            isGetResult = True
            break
        #print(row[0])
        counter += 1
    if(isGetResult == False):
        firstRow = sheet[1]
        for cell in firstRow:
            print(cell.value)
        result = firstRow[1].value
        monthExp = firstRow[2].value
        yearExp = firstRow[3].value

    return jsonify({
        "result": result,
        "month": monthExp,
        "year": yearExp,
        })